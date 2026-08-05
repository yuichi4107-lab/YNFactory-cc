"""
Character × Genre matrix image generator.

Usage:
  python generate_matrix.py --list
  python generate_matrix.py --character 25f --reference
  python generate_matrix.py --character 25f --genres
  python generate_matrix.py --character 25f --genres --only 1,3,5
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import time
from pathlib import Path

from openpyxl import load_workbook
from google import genai
from google.genai import types

HERE = Path(__file__).resolve().parent
XLSX = HERE / "character_settings_10people.xlsx"
OUT_ROOT = HERE / "output"
REF_DIR = OUT_ROOT / "references"
MATRIX_DIR = OUT_ROOT / "matrix"

MODEL = "gemini-2.5-flash-image"

AGE_KEYS = ["18", "25", "40", "55", "70"]
GENDER_KEYS = {"f": ("女性", 1), "m": ("男性", 2)}  # col idx in openpyxl (1-based)

# 採用する13ジャンル（人物描写が明確に変化するもの）。
# --only を明示指定した場合は除外対象も含めて生成可能。
ADOPTED_GENRE_INDICES: set[int] = {1, 2, 4, 6, 7, 8, 9, 10, 11, 12, 13, 16, 19}


def load_characters() -> dict[str, dict]:
    """Return dict keyed by '18f', '18m', ... with character text."""
    wb = load_workbook(XLSX, data_only=True)
    sh = wb["キャラクター設定"]
    out: dict[str, dict] = {}
    for row_idx, age_key in enumerate(AGE_KEYS, start=2):
        for gender_code, (gender_label, col) in GENDER_KEYS.items():
            text = sh.cell(row=row_idx, column=col + 1).value
            name_match = re.search(r"名前候補[：:]\s*(\S+)", text or "")
            name = name_match.group(1) if name_match else f"{age_key}{gender_label}"
            out[f"{age_key}{gender_code}"] = {
                "key": f"{age_key}{gender_code}",
                "age": age_key,
                "gender": gender_label,
                "name": name,
                "text": text.strip() if text else "",
                "slug": f"{age_key}歳_{gender_label}_{name}",
            }
    return out


def load_style_matrix() -> dict:
    wb = load_workbook(XLSX, data_only=True)
    sh = wb["シート1"]
    common_rules = []
    for row in range(5, 9):  # rows 5-8: 補足 bullets
        v = sh.cell(row=row, column=1).value
        if v:
            common_rules.append(v.strip())
    genres = []
    for col in range(1, sh.max_column + 1):
        genre = sh.cell(row=11, column=col).value
        draw = sh.cell(row=12, column=col).value
        tone = sh.cell(row=13, column=col).value
        line = sh.cell(row=14, column=col).value
        effect = sh.cell(row=15, column=col).value
        if not genre:
            continue
        # genre value looks like "- ジャンル: 恋愛に最適化した統一スタイル"
        g_name_match = re.search(r"ジャンル:\s*([^に]+)", genre)
        g_name = g_name_match.group(1).strip() if g_name_match else f"genre{col}"
        genres.append({
            "index": col,
            "name": g_name,
            "block": "\n".join(x for x in [genre, draw, tone, line, effect] if x),
        })
    return {"common_rules": common_rules, "genres": genres}


def sanitize(name: str) -> str:
    return re.sub(r"[^\w\-一-龥ぁ-んァ-ヶー]", "_", name)


def strip_props(text: str) -> str:
    """Remove the '小物：...' line from character text to avoid prop contradiction."""
    lines = [ln for ln in text.splitlines() if not re.match(r"\s*小物[：:]", ln)]
    return "\n".join(lines)


def soften_genre_block(block: str, genre_name: str) -> str:
    """Strip gore/violence phrases that trigger Gemini's IMAGE_SAFETY filter."""
    if "ホラー" in genre_name or "サスペンス" in genre_name:
        block = block.replace("血しぶき", "緊張感のある構図")
        block = block.replace("リアルな恐怖描写", "雰囲気による恐怖演出")
        block = block.replace("恐怖を煽る構図", "不穏で静かな構図")
    return block


# 背景に頼りがちなジャンルに「人物描写そのもので表現する画風」を強制指定。
# key は genre['name'] に含まれるキーワードで前方一致判定。
GENRE_STYLE_OVERRIDES: dict[str, str] = {
    "ビジネス": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: フラットベクターイラスト調（Adobe Stock のコーポレート系ベクター風）\n"
        "- 線画: 太めで均一なクリアな輪郭線、ディテールはシンプル化\n"
        "- 塗り: 完全フラット、陰影は極小（あっても1段階のみ）\n"
        "- 色調: ネイビー・グレー・白を基調に、アクセント1色（青またはオレンジ）\n"
        "- 背景: 完全に白または非常に薄いグレー単色（オフィス等の場面は描かない）"
    ),
    "哲学": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: モノクロ銅版画・エッチング風、思索的で静謐\n"
        "- 線画: 細い線の集積（クロスハッチング）で陰影を表現、エッジは硬質\n"
        "- 塗り: グレースケール主体、ハイライトと深い黒のコントラスト\n"
        "- 色調: モノクロ〜セピア、彩度は極力落とす\n"
        "- 背景: 無地、または古紙・羊皮紙風の薄いテクスチャのみ"
    ),
    "解説": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: 教育・ガイドブック風のポップなフラットイラスト\n"
        "- 線画: 太めで明快な輪郭、親しみやすいデフォルメ\n"
        "- 塗り: フラットで明るい原色系（青・黄・緑・赤）、陰影ほぼなし\n"
        "- 色調: 明度高く彩度中程度、視認性重視\n"
        "- 背景: 完全に白または単色、装飾は禁止"
    ),
    "投資": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: シャープでスタイリッシュなフラットベクターイラスト、高級感のある都市的デザイン\n"
        "- 線画: エッジの効いた細めの線、シンプル化された形\n"
        "- 塗り: フラット〜グラデーション、メタリックな光沢の表現\n"
        "- 色調: 深紺・金・銀・白、アクセントに緑（上昇色）\n"
        "- 背景: 単色またはシンプルなグラデーションのみ（グラフや数字は描かない）"
    ),
    "副業": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: カジュアルでモダンなフラットイラスト（SNS広告・LP挿絵風）\n"
        "- 線画: 柔らかく丸みのある輪郭、デフォルメされた人物\n"
        "- 塗り: フラット、ほんのり陰影、親しみやすい\n"
        "- 色調: オレンジ・明るい青・黄色・白、ポップで活発\n"
        "- 背景: 白または淡い単色、装飾禁止"
    ),
    "趣味": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: 絵本・水彩画調、手描きの温もり\n"
        "- 線画: 柔らかく丸みのある線、わずかにラフな筆跡\n"
        "- 塗り: 水彩の滲みと透明感、境界が柔らかい\n"
        "- 色調: 温かみのあるパステル〜中彩度、優しい配色\n"
        "- 背景: 水彩紙のテクスチャまたは白、余白を活かす"
    ),
    "論文": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: ペン画・インクドローイング調、エッチング風の理知的表現\n"
        "- 線画: 細密なクロスハッチングで陰影、硬質で正確な線\n"
        "- 塗り: モノクロまたはセピア単色、ハイライトは紙の白\n"
        "- 色調: モノトーン、アカデミックな抑制感\n"
        "- 背景: 無地の白または淡いクリーム色（数式・グラフ等は描かない）"
    ),
    "恋愛": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: 少女漫画風の繊細で柔らかなイラスト\n"
        "- 線画: 細く繊細、まつげや髪の毛流れを丁寧に\n"
        "- 塗り: ふんわりした水彩系グラデーション、肌に血色\n"
        "- 色調: パステル（ピンク・クリーム・水色）、キラキラのハイライト\n"
        "- 背景: 淡い花柄・光の粒程度に抑え、主役はあくまで人物"
    ),
    "ミステリー": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: 青年漫画誌風のリアル寄り描写\n"
        "- 線画: 細密で硬質、ハッチングで重厚な陰影\n"
        "- 塗り: モノトーン主体に深い青や赤の差し色、コントラスト強め\n"
        "- 色調: 暗めの落ち着いたトーン、夜を思わせる青味\n"
        "- 背景: 無地またはわずかな影のグラデーションのみ"
    ),
    "日常コメディ": (
        "【作画スタイル（固定指定）】\n"
        "- 画風: 4コマ漫画風のコミカルデフォルメ\n"
        "- 線画: 太めで丸みがあり、ラフで楽しげ\n"
        "- 塗り: フラットで明快、カラフル\n"
        "- 色調: ビビッドでポップ、明るい原色\n"
        "- 背景: 白または単色、汗マークや星など簡易な記号のみ"
    ),
}


def resolve_style_block(genre: dict) -> str:
    """Return override style if genre has one, else the softened Excel block."""
    for key, override in GENRE_STYLE_OVERRIDES.items():
        if key in genre["name"]:
            return override
    return soften_genre_block(genre["block"], genre["name"])


def client_from_env() -> genai.Client:
    key = os.environ.get("GEMINI_API_KEY")
    if not key:
        print("ERROR: set GEMINI_API_KEY environment variable.", file=sys.stderr)
        sys.exit(1)
    return genai.Client(api_key=key)


def save_images(response, dest: Path) -> tuple[list[Path], str | None]:
    """Returns (saved paths, block_reason or None)."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    idx = 0
    block_reason: str | None = None
    for cand in response.candidates or []:
        if cand.content is None or not cand.content.parts:
            fr = getattr(cand, "finish_reason", None)
            block_reason = str(fr) if fr else "empty_response"
            continue
        for part in cand.content.parts:
            if getattr(part, "inline_data", None) and part.inline_data.data:
                suffix = ".png"
                if idx == 0:
                    p = dest.with_suffix(suffix)
                else:
                    p = dest.with_name(dest.stem + f"_{idx}").with_suffix(suffix)
                p.write_bytes(part.inline_data.data)
                paths.append(p)
                idx += 1
    return paths, block_reason


def build_ref_prompt(char: dict, rules: list[str]) -> str:
    age = int(char["age"])
    age_note = ""
    if age >= 40:
        age_note = (
            f"\n【年齢表現 最重要】このキャラクターは **{age}歳** です。"
            f"年齢を視覚的にはっきりと反映させてください。若返らせないこと。具体的には：\n"
            f"- 顔立ち：目尻や口元に自然な年齢感（軽いシワ、肌の落ち着き）\n"
            f"- 髪色：設定指定通り（白髪混じり・グレイッシュ・白髪など）を省略せず明確に描く\n"
            f"- 表情・佇まい：年相応の落ち着きと経験を感じる雰囲気\n"
            f"- 体つきや所作：若者風に描かないこと\n"
        )
    return (
        "日本の漫画/アニメ風のキャラクター全身立ち姿イラスト。"
        "キャラクターリファレンスとして使用するため、顔立ち・髪型・服装・体型が明確に分かる画像にしてください。\n\n"
        "【キャラクター設定】\n" + strip_props(char["text"]) + "\n"
        + age_note +
        "\n【作画指示】\n"
        "- 作画スタイル: 現代的でクリアなアニメ調、細く滑らかな線画、やわらかな陰影\n"
        "- 色調: ナチュラルで清潔感のある色味、フラットで明るい背景\n"
        "- 背景: 無地の薄いグレー～白のプレーン背景\n"
        "- ポーズ: 正面向きの全身立ち姿、両手は自然に下ろす\n"
        "- 表情: 穏やかで柔らかい表情\n\n"
        "【重要】手には何も持たせないでください。バッグ・カバン・スマホ・本・書類・傘などの小道具は一切描画しないこと。両手は体の横に自然に下ろす、もしくは前で軽く組む。\n\n"
        "【共通ルール】\n" + "\n".join(rules) + "\n\n"
        "画像アスペクト比: 9:16 (縦長)"
    )


def build_genre_prompt(char: dict, rules: list[str], genre: dict) -> str:
    softened_block = resolve_style_block(genre)
    safety_note = ""
    if "ホラー" in genre["name"] or "サスペンス" in genre["name"]:
        safety_note = (
            "\n【表現の制約】グロテスクな血・暴力・残酷な描写は避けること。"
            "キャラクター本人は穏やかに立っている状態を保つこと。\n"
        )
    return (
        "添付した参照画像と**同一人物**を、まったく異なるジャンルの漫画家・イラストレーターが描いた別バージョンとして生成してください。\n\n"
        "【最重要方針】\n"
        "作風の表現は **人物の描写そのもの** で行うこと。背景はほぼ無地、または最小限の抽象的要素に留める。\n"
        "画風の差の 80% 以上を『キャラクターの絵柄（線・塗り・陰影・タッチ・色・服装の描き方）』で表現し、背景は 20% 未満の補助要素に抑えること。\n"
        "例：ビジネスジャンルであってもオフィス背景に頼らず、人物の線画の清潔さ・塗りのフラットさ・配色の知的さで表現する。学術ジャンルであっても数式の装飾に頼らず、人物のクールで抑制された描き方で表現する。\n\n"
        "【同一人物として保つ核（変えない）】\n"
        "- 顔立ちの特徴（目の形・鼻・口・顔の輪郭）\n"
        "- 髪型と髪色\n"
        "- 年齢感・性別\n"
        "- 体型・身長感\n\n"
        "【ジャンルごとに大きく変化させる要素（人物描写そのもの）】\n"
        "- 線画：太さ・荒さ・繊細さ・ハッチングの有無（例：劇画なら太く荒い線と陰のハッチング、水墨画なら筆の勢いのある黒線、コメディなら丸く柔らかい線）\n"
        "- 塗り：フラット / 厚塗り / セル調 / 水彩風 / 版画風 / ドット絵風 / グラデーション主体 など\n"
        "- 陰影：強いコントラスト / フラットで陰影なし / リアルな立体感 / アニメセル調の2段階 など\n"
        "- 色調・カラーパレット：指定に従い、肌・髪・服の色の扱い自体を変える\n"
        "- 服装：ジャンルに合わせて変えて良い（時代劇→着物、SF→近未来服、ファンタジー→冒険装束、など）。服のディテール・質感の描き方もジャンル画風に合わせる\n"
        "- 表情・ポーズ・佇まい：ジャンルの空気感に合わせる\n"
        "- 全体の絵柄：現代アニメ調 / 青年誌劇画 / 水墨画 / 浮世絵 / 西洋古典画 / 絵本調 / 3Dレンダリング風 / レトロアニメ 等、ジャンルに応じた画風で描く\n\n"
        "【背景の扱い】\n"
        "背景は無地（プレーン）または作風に合わせた単色・抽象的テクスチャ程度に留めること。建物・風景・小道具・文字・グラフ等の描き込みで作風を説明しない。作風はキャラの絵から伝わるようにする。\n\n"
        "【キャラクター設定（人物認識の参考）】\n" + strip_props(char["text"]) + "\n\n"
        "【このジャンルの作画指定】\n" + softened_block + "\n"
        + safety_note +
        "\n【重要】手には何も持たせないでください。バッグ・スマホ・小物類は一切描画しないこと。\n\n"
        "【共通ルール】\n" + "\n".join(rules) + "\n\n"
        "画像アスペクト比: 9:16 (縦長)"
    )


def generate_reference(char_key: str):
    chars = load_characters()
    styles = load_style_matrix()
    if char_key not in chars:
        print(f"unknown character: {char_key}. use --list to see options.")
        sys.exit(1)
    char = chars[char_key]
    client = client_from_env()
    prompt = build_ref_prompt(char, styles["common_rules"])
    print(f"[ref] generating reference for {char['slug']} ...")
    resp = client.models.generate_content(
        model=MODEL,
        contents=[prompt],
    )
    dest = REF_DIR / f"{sanitize(char['slug'])}.png"
    paths, block = save_images(resp, dest)
    if not paths:
        print(f"  no image returned (reason: {block}).")
        sys.exit(2)
    for p in paths:
        print(f"  saved: {p}")


def generate_genres(char_key: str, only: list[int] | None = None):
    chars = load_characters()
    styles = load_style_matrix()
    if char_key not in chars:
        print(f"unknown character: {char_key}")
        sys.exit(1)
    char = chars[char_key]
    ref_path = REF_DIR / f"{sanitize(char['slug'])}.png"
    if not ref_path.exists():
        print(f"reference not found: {ref_path}. run with --reference first.")
        sys.exit(1)
    ref_bytes = ref_path.read_bytes()
    client = client_from_env()
    target_dir = MATRIX_DIR / sanitize(char["slug"])
    target_dir.mkdir(parents=True, exist_ok=True)
    genres = styles["genres"]
    if only:
        genres = [g for g in genres if g["index"] in set(only)]
    else:
        genres = [g for g in genres if g["index"] in ADOPTED_GENRE_INDICES]
    for g in genres:
        prompt = build_genre_prompt(char, styles["common_rules"], g)
        fname = f"{g['index']:02d}_{sanitize(g['name'])}.png"
        dest = target_dir / fname
        if dest.exists():
            print(f"  skip (exists): {dest.name}")
            continue
        print(f"[genre {g['index']:02d}] {g['name']} → generating ...")
        try:
            resp = client.models.generate_content(
                model=MODEL,
                contents=[
                    types.Part.from_bytes(data=ref_bytes, mime_type="image/png"),
                    prompt,
                ],
            )
        except Exception as e:
            print(f"  ERROR: {e}")
            time.sleep(3)
            continue
        paths, block = save_images(resp, dest)
        if not paths:
            print(f"  SKIP genre {g['name']} (reason: {block})")
        else:
            for p in paths:
                print(f"  saved: {p.name}")
        time.sleep(1)  # gentle rate limit


def cmd_list():
    chars = load_characters()
    print("Available characters:")
    for k, c in chars.items():
        print(f"  {k:>4}  {c['slug']}")
    styles = load_style_matrix()
    print("\nGenres (✓=採用 / ·=除外):")
    for g in styles["genres"]:
        mark = "✓" if g["index"] in ADOPTED_GENRE_INDICES else "·"
        print(f"  {mark} {g['index']:>2}  {g['name']}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true")
    ap.add_argument("--character", help="e.g. 25f, 18m, 40f")
    ap.add_argument("--reference", action="store_true")
    ap.add_argument("--genres", action="store_true")
    ap.add_argument("--only", help="comma-separated genre indexes (1-20)")
    args = ap.parse_args()

    if args.list:
        cmd_list()
        return

    if not args.character:
        ap.error("--character is required (e.g. --character 25f). use --list to see.")

    if args.reference:
        generate_reference(args.character)
    elif args.genres:
        only = [int(x) for x in args.only.split(",")] if args.only else None
        generate_genres(args.character, only)
    else:
        ap.error("pass --reference or --genres.")


if __name__ == "__main__":
    main()
