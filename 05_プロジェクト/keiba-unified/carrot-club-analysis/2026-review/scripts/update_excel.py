# -*- coding: utf-8 -*-
import json, openpyxl
from openpyxl.styles import Font
from openpyxl.comments import Comment

FONT = Font(name='ＭＳ ゴシック', size=11)
FONT_B = Font(name='ＭＳ ゴシック', size=11, bold=True)
F = 'キャロット2026検討.xlsx'
wb = openpyxl.load_workbook(F)

# ---------- 1) 募集馬一覧に列追加 ----------
ws = wb['募集馬一覧']
old_notes = [ws['M1'].value, ws['M2'].value]
ws['M1'] = None; ws['M2'] = None
dams = json.load(open('dams2026.json', encoding='utf-8'))
def dam_info(name):
    d = dams.get(name.replace('（外）','').replace('の25','')) or {}
    born = d.get('dam_born'); age = d.get('dam_age_2025'); noff = d.get('n_offspring_reg')
    if born and not (1995 <= born <= 2020):
        age = None
    return age, noff
hdrs = {'L':'母年齢','M':'登録産駒数','N':'母年齢+産駒数','O':'馬体重(8月)','P':'基準スコア'}
for col, h in hdrs.items():
    c = ws[f'{col}1']; c.value = h; c.font = FONT
for r in range(2, 95):
    name = ws[f'C{r}'].value
    age, noff = dam_info(name or '')
    ws[f'L{r}'] = age; ws[f'M{r}'] = noff
    ws[f'N{r}'] = f'=IF(OR(L{r}="",M{r}=""),"",L{r}+M{r})'
    ws[f'P{r}'] = (f'=IF(F{r}="牡",1,0)+IF(OR(MONTH(H{r})=3,MONTH(H{r})=4),1,0)'
                   f'+IF(I{r}="ノーザンＦ",1,0)+IF(AND(L{r}>=8,L{r}<=11),1,0)'
                   f'+IF(AND(J{r}>=2500,J{r}<6000),1,0)+IF(O{r}>=430,1,0)')
    for col in 'LMNOP':
        ws[f'{col}{r}'].font = FONT
ws['R1'] = old_notes[0]; ws['R2'] = old_notes[1]
ws['R3'] = '※母年齢=2025年産駒時の年齢・登録産駒数=netkeiba登録ベースの参考値(未登録の産駒は数えられていません)。空欄は海外繁殖等で特定できなかった母。'
ws['R4'] = '※基準スコアは「検討基準」シート参照。O列(馬体重)とJ列(募集総額)は8/6発表後に入力すると自動加点されます。'
for r in range(1,5): ws[f'R{r}'].font = FONT
ws['L1'].comment = Comment('母年齢と登録産駒数はnetkeibaから自動取得した参考値。重要な馬はカタログで要確認。', 'Claude')
for col, w in {'L':8.6,'M':11,'N':13,'O':11.7,'P':10.5}.items():
    ws.column_dimensions[col].width = w

# ---------- 2) 検討基準シート ----------
if '検討基準' in wb.sheetnames: del wb['検討基準']
ws2 = wb.create_sheet('検討基準', 1)
rows = [
['キャロット出資 検討基準（2020〜2022年募集・全276頭のデビュー後成績分析より）'],
[''],
['全体基準値: 勝ち上がり率65% / 賞金が募集総額を超えた馬24% / 重賞勝ち馬16頭'],
['※賞金=中央+地方の獲得賞金(2026/7時点)。回収率は進上金・維持費を含まない「賞金/募集総額」の倍率。'],
[''],
['◆スコア基準(各1点・計6点)', '条件', '根拠(勝ち上がり率/回収≥1率)'],
['1. 性別', '牡馬', '牡74%・30% vs メス56%・20%。重賞16頭中13頭が牡'],
['2. 生まれ月', '3〜4月生まれ', '3月74%・29%/4月66%・29% vs 2月55%・19%'],
['3. 生産', 'ノーザンＦ', '重賞勝ち16頭すべてノーザンＦ系。その他牧場は重賞0'],
['4. 母年齢', '8〜11歳(産駒誕生時)', '67%・30%で全帯域のピーク。7歳以下62%/16歳+61%'],
['5. 募集総額', '2500〜5999万円', '2500-3999万が最良(71%・32%)。2500万未満55%/8000万+は回収率最低'],
['6. 馬体重', '募集時430kg以上', '430kg未満51%(牝は42%)/430-459は74%。※牡は420kg台でも大物例あり(シックスペンス・ドゥレッツァ)'],
[''],
['◆スコア別の過去実績', '', ''],
['スコア5〜6 (56頭)', '勝ち上がり86%・回収≥1が38%・重賞6頭', '中央値ベースの回収も全体の2倍超'],
['スコア4 (75頭)', '勝ち上がり67%・回収≥1が28%・重賞8頭', ''],
['スコア3以下 (145頭)', '勝ち上がり55%・回収≥1が17%・重賞2頭', '特にスコア≤2(67頭)は勝ち上がり52%'],
[''],
['◆父系の傾向(2020-22実績・今年もいる父のみ)', '', ''],
['キズナ', '勝ち上がり56%だが回収≥1が33%・重賞3頭', '当たれば大きい(一発型)'],
['モーリス', '勝ち上がり76%と高いが平均回収0.55', '手堅いが上値が重い'],
['ロードカナロア', '勝ち上がり67%・平均回収0.50', '価格が高くなりがちで妙味薄'],
['エピファネイア', '勝ち上がり62%・重賞0', '同上'],
['ドレフォン', '勝ち上がり64%・回収も平均以下', ''],
['リアルスティール', '勝ち上がり88%・回収≥1が38%', '少数だが優秀'],
['キタサンブラック', '勝ち上がり57%・重賞2頭(一発型)', 'イクイノックスの父。妙味あり'],
[''],
['◆その他の知見', '', ''],
['・関東/関西の差はほぼなし(回収≥1は関西36%とやや上)'],
['・「母年齢+産駒数」はご自身の仮説通り有効。合計10〜14がピーク(68%・28%)、9以下(若母・初仔)と20以上は割引'],
['・初仔は勝ち上がり62%と平均以下だがタスティエーラ(初仔・母6歳)の例外あり。母の質が高ければ許容'],
['・高額馬(8000万+)13頭で回収超えは2頭のみ。「良血高額=走る」は成立していない'],
['・メスの小柄(430kg未満)は57頭で勝ち上がり42%・回収中央値0.09と最も厳しい組み合わせ→回避推奨'],
['・2021年にご自身が印を付けた25頭: 大物2頭(タスティエーラ・ラヴェル)を的中。ただし率では印なしと同等で、'],
['  ドゥレッツァ/スキルヴィング/セラフィックコール(いずれも印なし牡・中価格帯)を取りこぼし→牡の中価格帯は広めに拾うのが吉'],
[''],
['◆注意', '', ''],
['・2022年募集世代(現4〜5歳)はまだ現役が多く、数値は今後上振れし得る'],
['・メスは引退後の繁殖価値が回収に含まれていないため、実質はやや過小評価'],
['・種牡馬の顔ぶれは毎年変わるため、父系の傾向は参考程度に'],
]
r = 1
for row in rows:
    for j, v in enumerate(row):
        if v:
            c = ws2.cell(row=r, column=j+1, value=v)
            c.font = FONT_B if (isinstance(v,str) and v.startswith('◆')) or r==1 else FONT
    r += 1
ws2.column_dimensions['A'].width = 34
ws2.column_dimensions['B'].width = 42
ws2.column_dimensions['C'].width = 55

# ---------- 3) 過去募集馬成績シート ----------
if '過去募集馬成績' in wb.sheetnames: del wb['過去募集馬成績']
ws3 = wb.create_sheet('過去募集馬成績')
data = json.load(open('dataset.json', encoding='utf-8'))
hdr = ['募集年','No','募集馬名','登録名','父','性別','生月','母年齢','産駒数','募集総額(万)','馬体重','通算成績','勝利数','中央賞金(万)','地方賞金(万)','賞金/募集額','重賞','主な勝ち鞍']
for j, h in enumerate(hdr, 1):
    c = ws3.cell(row=1, column=j, value=h); c.font = FONT_B
for i, x in enumerate(sorted(data, key=lambda x:(x['year'], int(x['no']))), start=2):
    vals = [int(x['year']), int(x['no']), x['name'], x['reg_name'], x['sire'], x['sex'],
            x.get('birth_month'), x.get('dam_age'), x.get('n_foals'), x['total_man'], x.get('weight'),
            x.get('record') or ('未登録' if not x['found'] else ''), x['wins'],
            x.get('prize_jra') or 0, x.get('prize_nar') or 0, round(x['ret'],2),
            '○' if x['graded'] else '', (x.get('main_wins') or '')[:60]]
    for j, v in enumerate(vals, 1):
        c = ws3.cell(row=i, column=j, value=v); c.font = FONT
        if j in (10,14,15): c.number_format = '#,##0'
        if j == 16: c.number_format = '0.00'
ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f'A1:R{len(data)+1}'
for col, w in {'C':26,'D':22,'E':16,'L':14,'R':40}.items():
    ws3.column_dimensions[col].width = w

wb.save(F)
print('saved')
