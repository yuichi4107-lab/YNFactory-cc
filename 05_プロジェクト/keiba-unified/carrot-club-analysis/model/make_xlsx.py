#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
2026年度募集の予測を配布用の Excel ブックに書き出す。

`forecast2026.py` が作る `data/forecast_2026.csv` と、募集馬リスト
`data/bosyu_2026.csv` を読んで、3シートのブックを組み立てる。

  1. 募集馬一覧＋予測 … 94頭の一覧に予測列を足したもの
  2. 母馬優先枠の予測 … 母馬優先対象57頭を確率順に並べたもの
  3. 読み方          … 手法・色の意味・精度・注意点

出力先は既定で `03_成果物/outputs/`（リポジトリの .gitignore 対象。
実体は Drive 側にあるので、ブック本体はリポジトリには入らない）。

使い方
------
  python3 model/make_xlsx.py
  python3 model/make_xlsx.py --out /path/to/dir
"""

from __future__ import annotations

import argparse
import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
ROOT = os.path.join(HERE, "..")
D_LIST = os.path.join(ROOT, "data", "bosyu_2026.csv")
D_FC = os.path.join(ROOT, "data", "forecast_2026.csv")
DEFAULT_OUT = os.path.join(ROOT, "..", "..", "..", "03_成果物", "outputs")
FILENAME = "キャロット2026_母馬優先枠_予測.xlsx"

HEAD_BASE = PatternFill("solid", fgColor="1F3864")
HEAD_ADD = PatternFill("solid", fgColor="2E5E3E")
FILL_HIGH = PatternFill("solid", fgColor="F8CBAD")     # 50%以上
FILL_MID = PatternFill("solid", fgColor="FFE699")      # 25〜50%
FILL_LOW = PatternFill("solid", fgColor="D9EAD3")      # 25%未満
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# 「母年齢」はクラブ確定一覧の定義（2025年産駒が生まれた時点の年齢）＝ 経過年数 t
BASE_COLS = ["No", "母馬優先", "募集馬名", "父", "母の父", "性別", "生月",
             "提供牧場", "総額(万円)", "一口(円)", "母年齢", "産駒数",
             "馬体重", "スコア", "地区", "予定厩舎"]
ADD_COLS = ["予測D(口)", "予測D下位10%", "予測D上位10%", "枠が埋まる確率",
            "馬体重補正", "補正なしの確率", "予測の根拠", "枠の口数"]

READING = [
    ["キャロット2026年度募集 母馬優先枠の予測"], [""],
    ["■ 何を予測しているか"],
    ["  母馬優先枠（総口数400口のうち200口／地方入厩予定馬は100口のうち50口）への申込口数 D と、"],
    ["  それが枠の口数に届いて『母馬優先者どうしの抽選』になる確率。"],
    ["  確率が高いほど、母馬優先権を持っていても落選しうるということ。"], [""],
    ["■ 色"],
    ["  赤＝50%以上（抽選になりそう）  黄＝25〜50%  緑＝25%未満（優先権を使えばまず通る）"], [""],
    ["■ モデル"],
    ["  log D(母馬,年) = m(母馬) − λ(年−2026) + ε"],
    ["   m … その母馬の水準（有資格者数と母系の格）。過去の観測から推定"],
    ["   λ … 年8.5%の減衰（退会4.5% ＋ 産次が進むことによる権利行使率の低下）"],
    ["   ε … その年の産駒しだいのブレ（σ=0.52）"], [""],
    ["■ 予測の根拠（列）"],
    ["  中間発表n年 … 締切前日の申込内訳から D を直接観測できた年数（いちばん強い）"],
    ["  抽選ランクn年 … D が枠の口数を超えたか否かだけが分かる年数（弱い）"],
    ["  母馬の馬齢のみ … 過去に産駒の募集が無く、馬齢からの事前分布だけ"], [""],
    ["■ 馬体重補正"],
    ["  2021・2022年度の92頭で、カタログ馬体重が母馬の馬齢と独立に効くことを確認した（z=+2.32）。"],
    ["  ただし効くのは軽い側だけ（年内平均−30kg以下の帯で27%、平均帯38%、+30kg以上でも42%）。"],
    ["  そこで『年内平均より軽い側にだけ減点』し、対象57頭の中で平均ゼロになるよう中心化している。"],
    ["  この項だけはバックテストできていないので、『補正なしの確率』も併記した。"], [""],
    ["■ 精度（バックテスト：その年より前のデータだけで過去を当てる）"],
    ["  2023年度 AUC 0.76 / 2024年度 0.79 / 2025年度 0.78"],
    ["  ブライアスコア 0.17〜0.18（全頭に一律34%と答えるモデルは0.224）"], [""],
    ["■ 全体"],
    ["  枠が埋まる見込み 22.5頭 / 57頭 = 40%（8割の幅で17〜28頭）"],
    ["  実測の推移 2021年度40% → 2022年度39% → 2023年度24% → 2024年度34% → 2025年度34%"], [""],
    ["■ 注意"],
    ["  ・年ごとの共通ショックは読めない（2023年度は24%まで落ちた）"],
    ["  ・母馬優先枠＝200口という配分はクラブ公開ページでは未確認（会員ブログとランク体系から）"],
    ["  ・募集馬リストは2026/8/1時点の確定一覧94頭。取り下げが出れば変わる"],
    ["  ・詳細は 05_プロジェクト/keiba-unified/carrot-club-analysis/docs/09_2026年度の予測.md"],
]


def load():
    with open(D_LIST, encoding="utf-8") as f:
        horses = list(csv.DictReader(f))
    with open(D_FC, encoding="utf-8") as f:
        fc = {r["母馬名"]: r for r in csv.DictReader(f)}
    return horses, fc


def basis(r: dict) -> str:
    if r["中間発表の観測年数"] != "0":
        return f"中間発表{r['中間発表の観測年数']}年"
    if r["抽選ランクの観測年数"] != "0":
        return f"抽選ランク{r['抽選ランクの観測年数']}年"
    return "母馬の馬齢のみ"


def fill_for(p: float) -> PatternFill:
    return FILL_HIGH if p >= 0.5 else (FILL_MID if p >= 0.25 else FILL_LOW)


def style_header(ws, n_base: int, n_total: int) -> None:
    for i in range(1, n_total + 1):
        c = ws.cell(1, i)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = HEAD_ADD if i > n_base else HEAD_BASE
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_all(wb: Workbook, horses: list[dict], fc: dict):
    ws = wb.active
    ws.title = "募集馬一覧＋予測"
    ws.append(BASE_COLS + ADD_COLS)
    style_header(ws, len(BASE_COLS), len(BASE_COLS) + len(ADD_COLS))
    for h in horses:
        f = fc.get(h["母馬名"]) if h["母馬優先対象"] == "1" else None
        vals = [int(h["No"]), "●" if h["母馬優先対象"] == "1" else "", h["募集馬名"],
                h["父"], h["母の父"], h["性別"], h["生月"], h["提供牧場"],
                int(h["募集総額_万円"].replace("万", "")), int(h["1口円"]),
                int(h["経過年数t"]), int(h["産駒数"]) if h["産駒数"] else "",
                float(h["馬体重"]) if h["馬体重"] else "",
                int(h["スコア"]) if h["スコア"] else "", h["入厩"], h["厩舎"]]
        if f:
            vals += [int(f["予測D_口"]), int(f["予測D_下位10%"]), int(f["予測D_上位10%"]),
                     float(f["枠が埋まる確率"]), float(f["馬体重補正_倍"]),
                     float(f["馬体重を使わない確率"]), basis(f), int(f["枠の口数"])]
        else:
            vals += ["", "", "", "", "", "", "母馬優先の対象外", ""]
        ws.append(vals)
        row = ws.max_row
        for i in range(1, len(BASE_COLS) + len(ADD_COLS) + 1):
            ws.cell(row, i).border = BORDER
        if f:
            p = float(f["枠が埋まる確率"])
            c = ws.cell(row, len(BASE_COLS) + 4)
            c.number_format = "0%"
            c.fill = fill_for(p)
            c.font = Font(bold=True)
            ws.cell(row, len(BASE_COLS) + 5).number_format = "0.00"
            ws.cell(row, len(BASE_COLS) + 6).number_format = "0%"
    ws.freeze_panes = "D2"
    set_widths(ws, [5, 8, 24, 20, 18, 6, 5, 11, 11, 10, 8, 7, 8, 6, 6, 14,
                    10, 12, 12, 12, 9, 11, 15, 9])


def sheet_forecast(wb: Workbook, horses: list[dict], fc: dict):
    ws = wb.create_sheet("母馬優先枠の予測")
    cols = ["順位", "母馬", "母馬の馬齢", "産駒数", "募集馬名", "父", "性",
            "総額(万円)", "馬体重", "枠の口数", "予測D(口)", "8割の幅(下)",
            "8割の幅(上)", "枠が埋まる確率", "補正なしの確率", "予測の根拠"]
    ws.append(cols)
    style_header(ws, 0, len(cols))
    by_dam = {h["母馬名"]: h for h in horses if h["母馬優先対象"] == "1"}
    for n, f in enumerate(sorted(fc.values(),
                                 key=lambda r: -float(r["枠が埋まる確率"])), 1):
        h = by_dam[f["母馬名"]]
        p = float(f["枠が埋まる確率"])
        ws.append([n, f["母馬名"], int(f["母馬の馬齢"]),
                   int(h["産駒数"]) if h["産駒数"] else "", h["募集馬名"], f["父"],
                   "牡" if f["性別"] == "牡馬" else "牝",
                   int(h["募集総額_万円"].replace("万", "")),
                   float(h["馬体重"]) if h["馬体重"] else "",
                   int(f["枠の口数"]), int(f["予測D_口"]), int(f["予測D_下位10%"]),
                   int(f["予測D_上位10%"]), p, float(f["馬体重を使わない確率"]),
                   basis(f)])
        row = ws.max_row
        for i in range(1, len(cols) + 1):
            ws.cell(row, i).border = BORDER
        c = ws.cell(row, 14)
        c.number_format = "0%"
        c.font = Font(bold=True)
        c.fill = fill_for(p)
        ws.cell(row, 15).number_format = "0%"
    ws.freeze_panes = "C2"
    set_widths(ws, [5, 18, 7, 7, 26, 20, 5, 11, 9, 9, 11, 11, 11, 13, 13, 16])


def sheet_reading(wb: Workbook):
    ws = wb.create_sheet("読み方")
    for line in READING:
        ws.append(line)
    ws.cell(1, 1).font = Font(bold=True, size=14)
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and v.startswith("■"):
            ws.cell(r, 1).font = Font(bold=True, size=11, color="1F3864")
    ws.column_dimensions["A"].width = 110


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=DEFAULT_OUT, help="出力先ディレクトリ")
    a = ap.parse_args()

    horses, fc = load()
    wb = Workbook()
    sheet_all(wb, horses, fc)
    sheet_forecast(wb, horses, fc)
    sheet_reading(wb)

    os.makedirs(a.out, exist_ok=True)
    path = os.path.join(a.out, FILENAME)
    wb.save(path)
    print(f"{path} を書き出しました（募集馬 {len(horses)}頭 / 母馬優先 {len(fc)}頭）")


if __name__ == "__main__":
    main()
