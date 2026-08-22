#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
2つのキャロット分析を1冊に統合した検討ブックを作る。

  A（親ディレクトリ）… 会員数・退会率の推計 → 母馬優先枠の申込口数 D と
                       「枠が埋まる確率」の予測（母馬優先対象57頭）
  B（2026-review）  … 過去3年276頭のデビュー後成績から導いた6点満点スコア
                       （確定94頭）

この2つは「当たるか（競争率）」と「走るか（見込み）」で別軸なので、
94頭を1行1頭で横に並べ、掛け合わせた判断区分を付ける。

  入力
    ../data/forecast_2026.csv   … A側の予測（母馬優先対象57頭）
    ../data/bosyu_2026.csv      … A側の募集馬リスト（94頭・母馬優先フラグ）
    ../2026-review/キャロット2026検討.xlsx … B側（スコア・厩舎相性・過去成績）
  出力
    ../2026-review/キャロット2026_統合検討.xlsx

使い方
------
  python3 scripts/build_unified.py            # 2026-review/ 直下で実行
  python3 scripts/build_unified.py --out DIR
"""

from __future__ import annotations

import argparse
import csv
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
REVIEW = os.path.join(HERE, "..")
ANALYSIS = os.path.join(REVIEW, "..")

SRC_XLSX = os.path.join(REVIEW, "キャロット2026検討.xlsx")
SRC_FC = os.path.join(ANALYSIS, "data", "forecast_2026.csv")
SRC_LIST = os.path.join(ANALYSIS, "data", "bosyu_2026.csv")
FILENAME = "キャロット2026_統合検討.xlsx"

# ---- 見た目 -------------------------------------------------------------
HEAD_B = PatternFill("solid", fgColor="1F3864")   # B側（走るか）由来の列
HEAD_A = PatternFill("solid", fgColor="2E5E3E")   # A側（当たるか）由来の列
HEAD_J = PatternFill("solid", fgColor="7F3F00")   # 統合して足した列
F_S = PatternFill("solid", fgColor="F4B183")      # S 最有力
F_A = PatternFill("solid", fgColor="FFE699")      # A 有力
F_B = PatternFill("solid", fgColor="D9EAD3")      # B 検討
F_AVOID = PatternFill("solid", fgColor="E6B8B7")  # 回避条件に該当
_thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

# スコア別の中央勝ち上がり率（2020〜2024年度募集・中央400口444頭の実績。検討基準シートより）
WINRATE = {4: "67%", 3: "57%", 2: "46%", 1: "37%", 0: "15%"}
# スコア別の「賞金が募集総額を超えた馬」の割合
RETRATE = {4: "37%", 3: "16%", 2: "14%", 1: "15%", 0: "4%"}


def load_b():
    """B側Excelの募集馬一覧（94頭）をキャッシュ値で読む。"""
    wb = load_workbook(SRC_XLSX, data_only=True)
    ws = wb["募集馬一覧"]
    hdr = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[2] is None:
            continue
        rows.append(dict(zip(hdr, r)))
    return wb, rows


def load_a():
    with open(SRC_FC, encoding="utf-8") as f:
        fc = {r["母馬名"]: r for r in csv.DictReader(f)}
    with open(SRC_LIST, encoding="utf-8") as f:
        lst = {r["募集馬名"]: r for r in csv.DictReader(f)}
    return fc, lst


def dam_of(name: str) -> str:
    """「ブランノワールの25」→「ブランノワール」"""
    return name.rsplit("の", 1)[0]


def ease(priority: bool, p):
    """取りやすさの一言。p は母馬優先枠が埋まる確率。"""
    if not priority:
        return "母馬優先者なし（400口すべて一般枠）"
    if p is None:
        return "母馬優先対象・予測なし"
    if p >= 0.50:
        return "母馬優先枠でも抽選が濃厚"
    if p >= 0.25:
        return "母馬優先枠が埋まる可能性あり"
    return "母馬優先枠は埋まりにくい"


def verdict(score, priority, p, sex, weight):
    """スコア（走るか）× 競争率（当たるか）の判断区分。

    走る側はスコア別の実績（4点=中央勝ち上がり67%・回収≥1が37% / 3点=57%・16% /
    2点=46%・14% / 1点=37%・15% / 0点=15%・4%）、
    当たる側は母馬優先枠が埋まる確率。母馬優先対象外は枠の食い合いがないので
    「取りやすい」側に置く。
    """
    s = score or 0
    tight = bool(priority) and p is not None and p >= 0.50
    if s >= 4:
        return "S 最有力" if not tight else "A 有力（人気）"
    if s == 3:
        return "A 有力" if not tight else "B 検討"
    if s == 2:
        return "C 見送り寄り"
    return "D 回避"


def avoid_flag(sex, weight):
    """検討基準シートの明示的な回避条件：メスの420kg未満。"""
    if sex and "メ" in str(sex) and weight not in (None, "") and float(weight) < 420:
        return "メス420kg未満（中央勝ち上がり29%・回収≥1が6%）"
    return ""


def style_header(ws, row, fills):
    for i, fill in enumerate(fills, start=1):
        c = ws.cell(row, i)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- 1. 統合判断表 -------------------------------------------------------
COLS = [
    ("No", HEAD_B, 5), ("募集馬名", HEAD_B, 20), ("父", HEAD_B, 15),
    ("母の父", HEAD_B, 15), ("性別", HEAD_B, 6), ("生月", HEAD_B, 6),
    ("提供牧場", HEAD_B, 11), ("総額(万円)", HEAD_B, 10), ("一口(円)", HEAD_B, 10),
    ("母年齢", HEAD_B, 7), ("産駒数", HEAD_B, 7), ("馬体重", HEAD_B, 8),
    ("地区", HEAD_B, 6), ("予定厩舎", HEAD_B, 11), ("厩舎相性", HEAD_B, 8),
    ("スコア", HEAD_B, 7), ("勝ち上がり実績", HEAD_B, 11), ("回収≥1実績", HEAD_B, 11),
    ("スコア内訳", HEAD_B, 26),
    ("母馬優先", HEAD_A, 8), ("予測D(口)", HEAD_A, 9), ("下位10%", HEAD_A, 8),
    ("上位10%", HEAD_A, 8), ("枠が埋まる確率", HEAD_A, 11), ("枠の口数", HEAD_A, 8),
    ("取りやすさ", HEAD_J, 26), ("判断区分", HEAD_J, 13), ("回避条件", HEAD_J, 30),
]


def sheet_unified(wb, brows, fc, alist):
    ws = wb.create_sheet("統合判断表")
    ws.append([c[0] for c in COLS])
    style_header(ws, 1, [c[1] for c in COLS])
    ws.freeze_panes = "C2"

    for b in brows:
        name = b["募集馬名"]
        a = alist.get(name, {})
        priority = a.get("母馬優先対象") == "1"
        f = fc.get(dam_of(name))
        p = float(f["枠が埋まる確率"]) if f else None
        score = b.get("スコア")
        sex = b.get("性別")
        wt = b.get("馬体重(カタログ)")
        bd = b.get("生年月日")
        row = [
            b.get("No"), name, b.get("父"), b.get("母の父"), sex,
            (bd.month if hasattr(bd, "month") else a.get("生月")),
            b.get("提供牧場"), b.get("総額(万円)"), b.get("一口(円)"),
            b.get("母年齢"), b.get("産駒数"), wt,
            b.get("地区"), b.get("予定厩舎"), b.get("厩舎相性"),
            score, WINRATE.get(score or 0), RETRATE.get(score or 0), b.get("スコア内訳"),
            "○" if priority else "",
            int(f["予測D_口"]) if f else None,
            int(f["予測D_下位10%"]) if f else None,
            int(f["予測D_上位10%"]) if f else None,
            p, int(f["枠の口数"]) if f else None,
            ease(priority, p),
            verdict(score, priority, p, sex, wt),
            avoid_flag(sex, wt),
        ]
        ws.append(row)

    # 列の位置は見出し名から引く（列を足しても壊れないように）
    idx = {c[0]: i for i, c in enumerate(COLS, start=1)}
    center = ["No", "性別", "生月", "母年齢", "産駒数", "馬体重", "地区", "スコア",
              "勝ち上がり実績", "回収≥1実績", "母馬優先", "予測D(口)", "下位10%", "上位10%",
              "枠が埋まる確率", "枠の口数"]
    center_idx = {idx[k] for k in center}
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(COLS) + 1):
            cell = ws.cell(r, c)
            cell.border = BORDER
            cell.font = Font(size=10)
            if c in center_idx:
                cell.alignment = Alignment(horizontal="center")
        pc = ws.cell(r, idx["枠が埋まる確率"])
        if pc.value is not None:
            pc.number_format = "0%"
        v = ws.cell(r, idx["判断区分"]).value
        fill = {"S 最有力": F_S, "A 有力": F_A, "A 有力（人気）": F_A,
                "B 検討": F_B}.get(v)
        if fill:
            for c in (idx["募集馬名"], idx["スコア"], idx["判断区分"]):
                ws.cell(r, c).fill = fill
        if ws.cell(r, idx["回避条件"]).value:
            ws.cell(r, idx["回避条件"]).fill = F_AVOID

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
    autosize(ws, [c[2] for c in COLS])
    ws.row_dimensions[1].height = 30
    return ws


# ---- 2. 読み方 -----------------------------------------------------------
READING = [
    ["キャロット2026年度募集 統合検討ブック"],
    [""],
    ["■ このブックは何か"],
    ["  別々に進めていた2つの分析を、94頭の一覧の上で1つに合わせたもの。"],
    ["   ・走るか … 過去5年（2020〜2024年度募集）467頭のデビュー後成績から導いた4点満点スコア"],
    ["   ・当たるか … 会員数・退会率の推計から出した、母馬優先枠の申込口数Dと枠が埋まる確率"],
    [""],
    ["■ 列の色分け（見出し行）"],
    ["  紺 … 成績分析（走るか）由来"],
    ["  緑 … 会員数・優先枠の推計（当たるか）由来"],
    ["  茶 … 2つを掛け合わせて足した列"],
    [""],
    ["■ スコア（0〜4点。各1点）"],
    ["  牡馬 / 募集総額2500万円以上 / 募集総額4000万円未満 / 募集時馬体重420kg以上"],
    ["  価格は下限と上限で効き方が違うので2本に割っている。"],
    ["  下限（2500万以上）は「走るか」に効き、上限（4000万未満）は「回収するか」にだけ効く。"],
    ["  過去5年の実績（中央400口444頭。勝ち上がりは中央で1勝以上）"],
    ["    4点(52頭) 中央勝ち上がり67%・回収≥1が37%・回収中央値0.7"],
    ["    3点(179頭) 中央勝ち上がり57%・回収≥1が16%・回収中央値0.26"],
    ["    2点(107頭) 中央勝ち上がり46%・回収≥1が14%・回収中央値0.19"],
    ["    1点(79頭) 中央勝ち上がり37%・回収≥1が15%・回収中央値0.17"],
    ["    0点(27頭) 中央勝ち上がり15%・回収≥1が4%・回収中央値0.03"],
    ["  2026/8に3年→5年へ広げて基準を作り直した。3〜4月生・ノーザンF生産・母8〜11歳は"],
    ["  5年では効かず（年度ダミーつきで有意でない）落とした。"],
    ["  その後、血統・母・馬体・価格・人気・厩舎・手法の7方向を並行で洗い直し、"],
    ["  検証を通ったのは「価格の下限を独立させる」1件だけだった。詳細は「検討基準」シート"],
    [""],
    ["■ 枠が埋まる確率"],
    ["  母馬優先枠（400口中200口／地方馬は100口中50口）への申込が枠を超え、"],
    ["  母馬優先者どうしの抽選になる確率。高いほど、優先権を持っていても落選しうる。"],
    ["  母馬優先対象外の37頭は、その母馬に優先権を持つ会員がいないため400口すべてが一般枠。"],
    ["  ※ 手法は親ディレクトリの docs/09_2026年度の予測.md"],
    [""],
    ["■ 判断区分"],
    ["  S 最有力    … スコア4 かつ 優先枠が埋まる確率50%未満（または母馬優先対象外）"],
    ["  A 有力      … スコア3 で競争が緩い、またはスコア4だが人気で競争が厳しい"],
    ["  B 検討      … スコア3 かつ 優先枠が埋まる確率50%以上"],
    ["  C 見送り寄り … スコア2"],
    ["  D 回避      … スコア1以下（過去5年106頭で中央勝ち上がり31%・回収が募集額を超えたのは12%）"],
    [""],
    ["■ 回避条件"],
    ["  メスの420kg未満は過去5年64頭で中央勝ち上がり29%・回収≥1が6%と最も厳しい組み合わせ。"],
    ["  判断区分と独立に、別列で印を付けている。"],
    [""],
    ["■ 注意"],
    ["  ・スコアは母数444頭の粗い層別であって、個別の馬の能力予測ではない。"],
    ["  ・当たり具合はAUC0.64（勝ち上がり）／0.61（回収）。でたらめが0.5なので、効きは穏やか。"],
    ["  ・2023年度募集世代は現4歳、2024年度募集世代は現3歳。成績は今後上振れし得る。"],
    ["  ・2026年度は募集価格が高く（中央値5000万）、上限基準に当てはまるのは94頭中20頭。"],
    ["  ・厩舎相性は各厩舎5〜17頭の小標本。参考程度に。"],
    ["  ・地方入厩予定の6頭は中央で走らないため、スコアの前提（中央400口444頭）の外にある。"],
    ["  ・枠が埋まる確率は推計であり、クラブの公表値ではない。前提は docs/04_出典.md。"],
    ["  ・再現: python3 scripts/build_unified.py"],
]


def sheet_reading(wb):
    ws = wb.create_sheet("読み方")
    for line in READING:
        ws.append(line)
    ws["A1"].font = Font(bold=True, size=14)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.startswith("■"):
            ws.cell(r, 1).font = Font(bold=True, size=11, color="1F3864")
    ws.column_dimensions["A"].width = 100
    return ws


# ---- 3. 既存シートの取り込み --------------------------------------------
def copy_sheet(src_wb, dst_wb, title, new_title=None, widths=None):
    """B側ブックのシートを値のまま写す。"""
    s = src_wb[title]
    d = dst_wb.create_sheet(new_title or title)
    for row in s.iter_rows(values_only=True):
        d.append(list(row))
    for c in d[1]:
        if c.value is not None:
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = HEAD_B
            c.alignment = Alignment(horizontal="center", wrap_text=True)
    if widths:
        autosize(d, widths)
    d.freeze_panes = "A2"
    return d


def sheet_forecast(wb, fc):
    """A側の予測を確率の高い順で。"""
    ws = wb.create_sheet("母馬優先枠の予測")
    rows = sorted(fc.values(), key=lambda r: -float(r["枠が埋まる確率"]))
    hdr = list(rows[0].keys())
    ws.append(hdr)
    style_header(ws, 1, [HEAD_A] * len(hdr))
    for r in rows:
        ws.append([r[k] for k in hdr])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(hdr) + 1):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).font = Font(size=10)
    autosize(ws, [16, 8, 8, 7, 16, 7, 11, 12, 8, 9, 10, 10, 12, 12, 12, 8, 9, 11, 13])
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=REVIEW)
    args = ap.parse_args()

    src_wb, brows = load_b()
    fc, alist = load_a()

    assert len(brows) == 94, f"募集馬が94頭でない: {len(brows)}"
    assert len(fc) == 57, f"予測が57頭でない: {len(fc)}"

    wb = Workbook()
    wb.remove(wb.active)
    sheet_unified(wb, brows, fc, alist)
    sheet_reading(wb)
    copy_sheet(src_wb, wb, "検討基準", widths=[110, 14, 14])
    sheet_forecast(wb, fc)
    copy_sheet(src_wb, wb, "過去募集馬成績",
               widths=[8, 5, 20, 24, 16, 6, 6, 8, 7, 11, 8, 16, 7, 12, 12, 11, 7, 22, 7, 26])

    os.makedirs(args.out, exist_ok=True)
    path = os.path.join(args.out, FILENAME)
    wb.save(path)
    print(f"書き出し: {path}")
    for ws in wb.worksheets:
        print(f"  {ws.title}: {ws.max_row}行 x {ws.max_column}列")


if __name__ == "__main__":
    main()
