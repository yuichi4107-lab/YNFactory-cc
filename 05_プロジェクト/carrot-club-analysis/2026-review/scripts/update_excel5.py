# -*- coding: utf-8 -*-
"""キャロット2026検討.xlsx を5年ベースに更新する。

  過去募集馬成績 … 2020〜2024年度募集の全頭（panel5.csv）に差し替え
  検討基準       … criteria.json の内容で書き直し
  募集馬一覧     … 新基準でスコアとスコア内訳を計算し直して値で書き込む

スコアは式ではなく値で入れる。openpyxlは式を計算しないので、式のまま置くと
このあとの build_unified.py がキャッシュ値を読めなくなるため。
"""
import csv
import io
import json
import os
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from scoring import load_criteria, score

BASE = os.path.dirname(os.path.abspath(__file__))
REVIEW = os.path.join(BASE, '..')
DS = os.path.join(REVIEW, 'datasets')
DATA = os.path.join(REVIEW, '..', 'data')
XLSX = os.path.join(REVIEW, 'キャロット2026検討.xlsx')

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
FONT = Font(name='ＭＳ ゴシック', size=11)
FONT_B = Font(name='ＭＳ ゴシック', size=11, bold=True)
HEAD_FILL = PatternFill('solid', fgColor='1F3864')
HEAD_FONT = Font(name='ＭＳ ゴシック', size=11, bold=True, color='FFFFFF')


def num(v, d=None):
    try:
        return float(str(v).replace(',', ''))
    except Exception:
        return d


def clear(ws):
    ws.delete_rows(1, ws.max_row)


def write_panel(wb, crit):
    rows = list(csv.DictReader(open(os.path.join(DS, 'panel5.csv'), encoding='utf-8-sig')))
    rs = json.load(open(os.path.join(DS, 'race_summary.json'), encoding='utf-8'))
    name = '過去募集馬成績'
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    clear(ws)
    hdr = ['募集年度', 'No', '募集馬名', '登録名', '父', '性別', '生月', '母年齢', '産駒数',
           '募集総額(万)', '馬体重', '体高', '胸囲', '管囲', '口数', '予定厩舎', '通算成績',
           '中央勝利', '地方勝利', '中央賞金(万)', '地方賞金(万)', '賞金/募集額', '重賞',
           '主な勝ち鞍', 'スコア', 'スコア内訳']
    ws.append(hdr)
    for c in ws[1]:
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(horizontal='center')
    for r in rows:
        key = r['year'] + '#' + r['no']
        s = rs.get(key) or {}
        row = dict(r)
        row['sex'] = r['sex']
        pts, detail, missing = score(row, crit)
        ws.append([
            int(r['year']), num(r['no']), r['name'], r['reg_name'], r['sire'], r['sex'],
            num(r['month']), num(r['dam_age']), num(r['n_foals']), num(r['total_man']),
            num(r['weight']), num(r['height']), num(r['girth']), num(r['cannon']),
            num(r['kuchi']), r['trainer_planned'] or r['trainer'],
            (f"{int(num(r['starts'], 0))}戦{int(num(r['wins'], 0))}勝" if r['starts'] else ''),
            s.get('jra_wins'), s.get('nar_wins'), num(r['prize_jra']), num(r['prize_nar']),
            round(num(r['ret'], 0), 2), int(num(r['graded'], 0)), r['main_wins'],
            pts if not missing else None, detail,
        ])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = FONT
    ws.freeze_panes = 'A2'
    print(f'  {name}: {len(rows)}頭')


def write_criteria(wb, crit):
    name = '検討基準'
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    clear(ws)
    for line in crit['sheet_lines']:
        ws.append(line)
    for row in ws.iter_rows():
        for c in row:
            c.font = FONT
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and (v.startswith('◆') or r == 1):
            ws.cell(r, 1).font = Font(name='ＭＳ ゴシック', size=11, bold=True, color='1F3864')
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    print(f'  {name}: {ws.max_row}行')


def trainer_marks():
    """5年ぶんの預託実績から、予定厩舎につける相性マークを決める。"""
    import pandas as pd

    from analyze5 import load
    df = load(central_only=True).dropna(subset=['win_jra'])
    t = df.groupby('trainer_key').agg(n=('win_jra', 'size'), win=('win_jra', 'mean'),
                                      ret1=('ret1', 'mean'), med=('ret', 'median'))
    marks = {}
    for name, v in t.iterrows():
        if v['n'] < 5:
            continue
        if v['ret1'] >= 0.30:
            marks[name] = ('◎', v)
        elif v['ret1'] >= 0.20 or v['win'] >= 0.65:
            marks[name] = ('○', v)
        elif v['ret1'] <= 0.05 and v['win'] <= 0.45:
            marks[name] = ('△', v)
    return marks


def rewrite_affinity(wb):
    ws = wb['募集馬一覧']
    hdr = {c.value: c.column for c in ws[1]}
    if '厩舎相性' not in hdr:
        return
    marks = trainer_marks()
    n = 0
    for r in range(2, ws.max_row + 1):
        tr = (ws.cell(r, hdr['予定厩舎']).value or '').replace(' ', '')
        m = marks.get(tr)
        ws.cell(r, hdr['厩舎相性'], m[0] if m else '').font = FONT
        if m:
            n += 1
    print(f'  厩舎相性: {n}頭に印（5頭以上預託のある厩舎のみ）')


def rescore_2026(wb, crit):
    ws = wb['募集馬一覧']
    hdr = {c.value: c.column for c in ws[1]}
    col_score, col_detail = hdr['スコア'], hdr['スコア内訳']
    n = 0
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, hdr['募集馬名']).value:
            continue
        row = {
            'sex': ws.cell(r, hdr['性別']).value,
            'farm': ws.cell(r, hdr['提供牧場']).value,
            'month': (ws.cell(r, hdr['生年月日']).value.month
                      if hasattr(ws.cell(r, hdr['生年月日']).value, 'month') else None),
            'dam_age': ws.cell(r, hdr['母年齢']).value,
            'total_man': ws.cell(r, hdr['総額(万円)']).value,
            'weight': ws.cell(r, hdr['馬体重(カタログ)']).value,
        }
        pts, detail, missing = score(row, crit)
        ws.cell(r, col_score, pts).font = FONT
        ws.cell(r, col_detail, detail).font = FONT
        n += 1
    print(f'  募集馬一覧: {n}頭を再採点')


def main():
    crit = load_criteria()
    wb = load_workbook(XLSX)
    print('更新:', os.path.basename(XLSX))
    write_panel(wb, crit)
    write_criteria(wb, crit)
    rescore_2026(wb, crit)
    rewrite_affinity(wb)
    wb.save(XLSX)
    print('保存しました')


if __name__ == '__main__':
    main()
